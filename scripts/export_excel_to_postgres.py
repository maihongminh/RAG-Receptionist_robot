#!/usr/bin/env python3
import argparse
import csv
import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import openpyxl


def safe_name(value, fallback):
    raw = str(value or "").strip().lower()
    raw = re.sub(r"[^a-z0-9_]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_")
    if not raw:
        raw = fallback
    if raw[0].isdigit():
        raw = f"c_{raw}"
    return raw


def unique_name(name, seen):
    if name not in seen:
        seen[name] = 1
        return name
    seen[name] += 1
    return f"{name}_{seen[name]}"


def quote_ident(identifier):
    return '"' + identifier.replace('"', '""') + '"'


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def build_sheet_mapping(workbook):
    seen_tables = {}
    mappings = []

    for sheet in workbook.worksheets:
        table_name = unique_name(safe_name(sheet.title, "sheet"), seen_tables)
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())

        headers = []
        seen_columns = {}
        for column_index in range(1, sheet.max_column + 1):
            original = first_row[column_index - 1] if column_index <= len(first_row) else None
            column_name = unique_name(
                safe_name(original, f"column_{column_index}"),
                seen_columns,
            )
            headers.append(
                {
                    "index": column_index,
                    "original": cell_to_text(original),
                    "name": column_name,
                }
            )

        mappings.append(
            {
                "sheet": sheet.title,
                "table": table_name,
                "rows": sheet.max_row,
                "cols": sheet.max_column,
                "columns": headers,
            }
        )

    return mappings


def write_csv_files(workbook, mappings, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)

    for mapping in mappings:
        sheet = workbook[mapping["sheet"]]
        csv_path = output_dir / f"{mapping['table']}.csv"
        column_names = ["_excel_row_number"] + [
            column["name"] for column in mapping["columns"]
        ]

        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(column_names)

            rows = sheet.iter_rows(
                min_row=2,
                max_row=sheet.max_row,
                max_col=sheet.max_column,
                values_only=True,
            )
            for row_number, excel_row in enumerate(rows, start=2):
                row = [row_number]
                has_value = any(value is not None for value in excel_row)
                row.extend(cell_to_text(value) for value in excel_row)

                if has_value:
                    writer.writerow(row)

        mapping["csv"] = str(csv_path)


def write_schema_sql(mappings, schema_name, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "-- Generated from clinic_full_export.xlsx.",
        "-- Raw import schema: every Excel cell is stored as TEXT to preserve source data.",
        f"DROP SCHEMA IF EXISTS {quote_ident(schema_name)} CASCADE;",
        f"CREATE SCHEMA {quote_ident(schema_name)};",
        "",
    ]

    for mapping in mappings:
        table = quote_ident(mapping["table"])
        columns = ['  "_excel_row_number" INTEGER NOT NULL']
        columns.extend(
            f"  {quote_ident(column['name'])} TEXT"
            for column in mapping["columns"]
        )

        lines.extend(
            [
                f"CREATE TABLE {quote_ident(schema_name)}.{table} (",
                ",\n".join(columns),
                ");",
                "",
            ]
        )

    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_load_sql(mappings, schema_name, csv_dir, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "-- Run from the project root with psql, for example:",
        "-- psql postgresql://USER:PASSWORD@HOST:PORT/DB_NAME -f db/import_all.sql",
        "",
    ]

    for mapping in mappings:
        table = quote_ident(mapping["table"])
        column_names = ["_excel_row_number"] + [
            column["name"] for column in mapping["columns"]
        ]
        columns_sql = ", ".join(quote_ident(column) for column in column_names)
        csv_path = (csv_dir / f"{mapping['table']}.csv").as_posix()
        csv_literal = csv_path.replace("'", "''")

        lines.append(
            f"\\copy {quote_ident(schema_name)}.{table} ({columns_sql}) "
            f"FROM '{csv_literal}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8');"
        )

    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_import_all_sql(schema_path, load_path, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "\\set ON_ERROR_STOP on",
        f"\\i {schema_path.as_posix()}",
        f"\\i {load_path.as_posix()}",
    ]
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_manifest(mappings, schema_name, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    manifest = {
        "schema": schema_name,
        "tables": mappings,
    }
    output_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main():
    parser = argparse.ArgumentParser(
        description="Export every Excel sheet to Postgres-ready CSV and SQL files."
    )
    parser.add_argument(
        "--excel",
        default="clinic_full_export.xlsx",
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--schema",
        default="robo_raw",
        help="Postgres schema name for raw imported tables.",
    )
    parser.add_argument(
        "--csv-dir",
        default="data/postgres_csv",
        help="Directory for generated CSV files.",
    )
    parser.add_argument(
        "--db-dir",
        default="db",
        help="Directory for generated SQL and manifest files.",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    csv_dir = Path(args.csv_dir)
    db_dir = Path(args.db_dir)

    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    mappings = build_sheet_mapping(workbook)

    write_csv_files(workbook, mappings, csv_dir)
    write_schema_sql(mappings, args.schema, db_dir / "schema.sql")
    write_load_sql(mappings, args.schema, csv_dir, db_dir / "load.sql")
    write_import_all_sql(
        Path("db/raw/schema.sql"),
        Path("db/raw/load.sql"),
        db_dir / "import_all.sql",
    )
    write_manifest(mappings, args.schema, db_dir / "manifest.json")

    print(
        json.dumps(
            {
                "excel": str(excel_path),
                "schema": args.schema,
                "sheet_count": len(mappings),
                "csv_dir": str(csv_dir),
                "sql": {
                    "schema": str(db_dir / "schema.sql"),
                    "load": str(db_dir / "load.sql"),
                    "import_all": str(db_dir / "import_all.sql"),
                    "manifest": str(db_dir / "manifest.json"),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
